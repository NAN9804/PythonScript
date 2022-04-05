import re
import sys
import openpyxl as xl

"""
本模块用于从Excel表格中提取约束管脚信息并生成.XDC文件

使用前修改sourcefile名称
Excel工作簿中用来存放约束信息的sheet需要命名为"xdc_sourcefile"
B列存放PROT_NAME，C列存放PACKAGE_PIN，D列存放IOSTANDARD
单元格中的内容需要规范，若一行中有一个空单元格，则此行将不被处理
"""

__author__ = 'NAN'

def xdc_generate(ws):
    """
    本函数从打开的工作表中提取约束信息并生成xdc文件
    """
    xdc_file = ''
    xdc_list = []
    i = 0

    maxrow = ws.max_row + 1
    for row in range(2,maxrow):
        col_B = 'B' + str(row)
        col_C = 'C' + str(row)
        col_D = 'D' + str(row)
        
        if ws[col_B].value!=None and ws[col_C].value!=None and ws[col_D].value!=None:
            portname = ws[col_B].value.strip()

            xdc_file = ('set_property PACKAGE_PIN ' + 
                        ws[col_C].value + 
                        ' [get_ports ' + 
                        portname + 
                        ']\n' + 
                        'set_property IOSTANDARD ' + 
                        ws[col_D].value + 
                        ' [get_ports ' + 
                        portname + 
                        ']\n')       
            xdc_list = xdc_list + [ xdc_file ]
        else:
            i = i + 1

    print('Have ' + str(i) + ' wrong type row(s)')
    xdc_file = '\n'.join(xdc_list)

    return xdc_file


if __name__ == '__main__':
    # sourcefile = sys.argv[1]
    sourcefile = 'NH7020_constr.xlsx' #文件名使用前修改
    try:
        wb = xl.load_workbook(sourcefile, read_only=True)
        ws = wb["xdc_sourcefile"]
    except FileNotFoundError:
        print('Sorry, the Excel does not exist!')
    except KeyError:
        print('Sorry, the sheet does not exist!')
    else:
        print('Open successfully!')

        xdc_file = xdc_generate(ws)

        #将生成的xdc文本内容写入文件
        with open('constr.xdc','w') as dest_file:
            dest_file.write(xdc_file)

        print('XDC file generated successfully!')
